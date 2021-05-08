#!/usr/bin/env python3
# -*-encoding: utf-8-*-

from routes.models import Shift, Schedule
from openpyxl import Workbook
from accounts.services import get_drivers
from django.utils.timezone import datetime, timedelta


def report_worked_hours(driver, date_from, date_to):
    shifts = Shift.objects.filter(driver=driver, date__range=[date_from, date_to])
    schedules = []
    res = []
    hours = 6 * len(shifts)
    for shift in shifts:
        schedules += [Schedule.objects.filter(shift=shift)]
        res.append(f'{shift.date}, {shift}')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'Report "worked hours"'
    ws['B1'] = f'from {date_from} to {date_to}'
    ws['C1'] = f'{driver.first_name} {driver.last_name}'
    ws['D1'] = f'total hours: {hours}'
    for i in range(len(res)):
        ws[f'A{i+2}'] = res[i]

    # wb.save(f'{driver}_{date_from}_{date_to}.xlsx')
    return wb


def report_worked_hours_all_drivers(date_from, date_to):

    wb = Workbook()

    drivers = get_drivers()
    for driver in drivers:
        wc = wb.create_sheet(f"Водій {driver.get_full_name()}")
        wc.title = f'Водій {driver.get_full_name()}'

        shifts = Shift.objects.filter(driver=driver, date__range=[date_from, date_to])
        schedules = []
        res = []
        hours = 6 * len(shifts)
        for shift in shifts:
            schedules += [Schedule.objects.filter(shift=shift)]
            res.append(f'{shift.date}, {shift}')

        wc['A1'] = f'Report "worked hours"'
        wc['B1'] = f'from {date_from} to {date_to}'
        wc['C1'] = f'{driver.get_full_name()}'
        wc['D1'] = f'total hours: {hours}'
        for i in range(len(res)):
            wc[f'A{i + 2}'] = res[i]
    return wb


def report_drivers_dest(date_from, date_to):
    wb = Workbook()

    drivers = get_drivers()
    for driver in drivers:
        wc = wb.create_sheet(f"Водій {driver.get_full_name()}")
        wc.title = f'Водій {driver.get_full_name()}'

        shifts = Shift.objects.filter(driver=driver, date__range=[date_from, date_to])
        wc['A1'] = f'Report "drivers dest"'
        wc['B1'] = f'from {date_from} to {date_to}'
        wc['C1'] = f'{driver.get_full_name()}'
        i = 0
        for shift in shifts:
            wc[f'A{i + 2}'] = str(shift)
            i += 1
    return wb


def report_kzot(driver, date_from, date_to):
    # shifts = Shift.objects.filter(driver=driver, date__range=[date_from, date_to])

    # get all weeks

    weeks = []
    date_from = datetime.strptime(date_from, '%Y-%m-%d')
    date_to = datetime.strptime(date_to, '%Y-%m-%d')

    cur_date = date_from

    while cur_date < date_to:
        monday_date = cur_date - timedelta(days=cur_date.weekday())
        sunday_date = monday_date + timedelta(days=6)

        weeks.append((monday_date, sunday_date))
        cur_date = cur_date + timedelta(days=7)

    violation = False
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'Report "KZOT"'
    ws['B1'] = f'from {date_from} to {date_to}'
    ws['C1'] = f'{driver.first_name} {driver.last_name}'
    i = 2
    for week in weeks:

        start, end = week
        shift_num = Shift.objects.filter(driver=driver, date__range=[start, end]).count()

        if shift_num > 7:
            violation = True
            ws[f'A{i}'] = f'week from {start.strftime("%Y-%m-%d")} to {end.strftime("%Y-%m-%d")}'
            ws[f'B{i}'] = f'worked hours {shift_num * 6}'
            i += 1

    if not violation:
        ws['A2'] = 'There were no violations'

    return wb


def report_kzot_all_drivers(date_from, date_to):

    wb = Workbook()

    date_from = datetime.strptime(date_from, '%Y-%m-%d')
    date_to = datetime.strptime(date_to, '%Y-%m-%d')

    weeks = []

    cur_date = date_from

    while cur_date < date_to:
        monday_date = cur_date - timedelta(days=cur_date.weekday())
        sunday_date = monday_date + timedelta(days=6)

        weeks.append((monday_date, sunday_date))
        cur_date = cur_date + timedelta(days=7)

    drivers = get_drivers()
    for driver in drivers:
        wc = wb.create_sheet(f"Водій {driver.get_full_name()}")
        wc.title = f'Водій {driver.get_full_name()}'

        violation = False
        wc['A1'] = f'Report "KZOT"'
        wc['B1'] = f'from {date_from} to {date_to}'
        wc['C1'] = f'{driver.first_name} {driver.last_name}'
        i = 2
        for week in weeks:

            start, end = week
            shift_num = Shift.objects.filter(driver=driver, date__range=[start, end]).count()

            if shift_num > 7:
                violation = True
                wc[f'A{i}'] = f'week from {start.strftime("%Y-%m-%d")} to {end.strftime("%Y-%m-%d")}'
                wc[f'B{i}'] = f'worked hours {shift_num * 6}'
                i += 1

        if not violation:
            wc['A2'] = 'There were no violations'

    return wb
